"""Analytics export service for various formats."""

import io
import json
import csv
from datetime import datetime
from typing import Dict, List, Any, Optional, Union
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from core.logging import logger


class AnalyticsExportService:
    """Service for exporting analytics data in various formats."""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._init_custom_styles()
        
    def _init_custom_styles(self):
        """Initialize custom PDF styles."""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#34495E'),
            spaceAfter=12
        ))
        
    async def export_to_csv(
        self,
        data: Dict[str, Any],
        filename: Optional[str] = None
    ) -> io.BytesIO:
        """
        Export analytics data to CSV format.
        
        Args:
            data: Analytics data dictionary
            filename: Optional filename
            
        Returns:
            BytesIO object containing CSV data
        """
        output = io.StringIO()
        
        # Handle different data structures
        if "overview" in data:
            # Dashboard export
            await self._write_dashboard_csv(output, data)
        elif "campaigns" in data:
            # Campaign export
            await self._write_campaign_csv(output, data)
        elif "users" in data:
            # User analytics export
            await self._write_user_csv(output, data)
        else:
            # Generic export
            await self._write_generic_csv(output, data)
            
        # Convert to bytes
        output.seek(0)
        bytes_output = io.BytesIO(output.getvalue().encode('utf-8'))
        bytes_output.seek(0)
        
        return bytes_output
        
    async def export_to_excel(
        self,
        data: Dict[str, Any],
        filename: Optional[str] = None
    ) -> io.BytesIO:
        """
        Export analytics data to Excel format with formatting and charts.
        
        Args:
            data: Analytics data dictionary
            filename: Optional filename
            
        Returns:
            BytesIO object containing Excel data
        """
        output = io.BytesIO()
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add sheets based on data structure
        if "overview" in data:
            await self._create_dashboard_excel(wb, data)
        if "phishing_stats" in data:
            await self._create_phishing_excel(wb, data["phishing_stats"])
        if "course_performance" in data:
            await self._create_course_excel(wb, data["course_performance"])
        if "user_activity" in data:
            await self._create_user_activity_excel(wb, data["user_activity"])
            
        # Save to BytesIO
        wb.save(output)
        output.seek(0)
        
        return output
        
    async def export_to_pdf(
        self,
        data: Dict[str, Any],
        filename: Optional[str] = None,
        company_name: Optional[str] = None
    ) -> io.BytesIO:
        """
        Export analytics data to PDF format with charts and formatting.
        
        Args:
            data: Analytics data dictionary
            filename: Optional filename
            company_name: Company name for report header
            
        Returns:
            BytesIO object containing PDF data
        """
        output = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            output,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Build content
        story = []
        
        # Title page
        story.append(Paragraph(
            f"{company_name or 'Company'} Security Analytics Report",
            self.styles['CustomTitle']
        ))
        story.append(Spacer(1, 0.5 * inch))
        story.append(Paragraph(
            f"Generated on {datetime.utcnow().strftime('%B %d, %Y')}",
            self.styles['Normal']
        ))
        story.append(PageBreak())
        
        # Executive summary
        if "overview" in data:
            story.extend(await self._create_executive_summary(data["overview"]))
            story.append(PageBreak())
            
        # Security score
        if "security_score" in data:
            story.extend(await self._create_security_score_section(data["security_score"]))
            story.append(PageBreak())
            
        # Risk assessment
        if "risk_assessment" in data:
            story.extend(await self._create_risk_assessment_section(data["risk_assessment"]))
            story.append(PageBreak())
            
        # Phishing statistics
        if "phishing_stats" in data:
            story.extend(await self._create_phishing_section(data["phishing_stats"]))
            story.append(PageBreak())
            
        # Course performance
        if "course_performance" in data:
            story.extend(await self._create_course_section(data["course_performance"]))
            story.append(PageBreak())
            
        # Recommendations
        if "recommendations" in data:
            story.extend(await self._create_recommendations_section(data["recommendations"]))
            
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        return output
        
    async def export_to_json(
        self,
        data: Dict[str, Any],
        pretty: bool = True
    ) -> io.BytesIO:
        """
        Export analytics data to JSON format.
        
        Args:
            data: Analytics data dictionary
            pretty: Whether to pretty-print JSON
            
        Returns:
            BytesIO object containing JSON data
        """
        output = io.BytesIO()
        
        # Add metadata
        export_data = {
            "metadata": {
                "exported_at": datetime.utcnow().isoformat(),
                "version": "1.0",
                "format": "json"
            },
            "data": data
        }
        
        # Write JSON
        if pretty:
            json_str = json.dumps(export_data, indent=2, default=str)
        else:
            json_str = json.dumps(export_data, default=str)
            
        output.write(json_str.encode('utf-8'))
        output.seek(0)
        
        return output
        
    # CSV Helper Methods
    
    async def _write_dashboard_csv(self, output: io.StringIO, data: Dict[str, Any]):
        """Write dashboard data to CSV."""
        writer = csv.writer(output)
        
        # Overview section
        writer.writerow(["Dashboard Overview"])
        writer.writerow(["Metric", "Value"])
        
        if "overview" in data:
            overview = data["overview"]
            writer.writerow(["Total Users", overview.get("total_users", 0)])
            writer.writerow(["Active Users", overview.get("active_users", 0)])
            writer.writerow(["Engagement Rate", f"{overview.get('engagement_rate', 0)}%"])
            writer.writerow(["Courses Completed", overview.get("courses_completed", 0)])
            writer.writerow(["Average Quiz Score", f"{overview.get('avg_quiz_score', 0)}%"])
            
        writer.writerow([])  # Empty row
        
        # User activity
        if "user_activity" in data and "most_active_users" in data["user_activity"]:
            writer.writerow(["Most Active Users"])
            writer.writerow(["Name", "Activity Count"])
            for user in data["user_activity"]["most_active_users"]:
                writer.writerow([user["name"], user["activity_count"]])
                
    async def _write_campaign_csv(self, output: io.StringIO, data: Dict[str, Any]):
        """Write campaign data to CSV."""
        writer = csv.writer(output)
        
        if "campaigns" in data:
            writer.writerow(["Campaign Performance"])
            writer.writerow(["Campaign Name", "Type", "Sent", "Opened", "Clicked", "Open Rate", "Click Rate"])
            
            for campaign in data["campaigns"]:
                writer.writerow([
                    campaign.get("name", ""),
                    campaign.get("type", ""),
                    campaign.get("sent", 0),
                    campaign.get("opened", 0),
                    campaign.get("clicked", 0),
                    f"{campaign.get('open_rate', 0)}%",
                    f"{campaign.get('click_rate', 0)}%"
                ])
                
    # Excel Helper Methods
    
    async def _create_dashboard_excel(self, wb: Workbook, data: Dict[str, Any]):
        """Create dashboard sheet in Excel."""
        ws = wb.create_sheet("Dashboard")
        
        # Header styling
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "Security Analytics Dashboard"
        ws["A1"].font = Font(bold=True, size=18)
        ws["A1"].alignment = header_alignment
        
        # Overview metrics
        if "overview" in data:
            overview = data["overview"]
            
            # Metrics headers
            metrics_row = 3
            ws[f"A{metrics_row}"] = "Metric"
            ws[f"B{metrics_row}"] = "Value"
            
            for col in ["A", "B"]:
                ws[f"{col}{metrics_row}"].font = header_font
                ws[f"{col}{metrics_row}"].fill = header_fill
                
            # Metrics data
            metrics_data = [
                ("Total Users", overview.get("total_users", 0)),
                ("Active Users", overview.get("active_users", 0)),
                ("Engagement Rate", f"{overview.get('engagement_rate', 0)}%"),
                ("Courses Completed", overview.get("courses_completed", 0)),
                ("Average Quiz Score", f"{overview.get('avg_quiz_score', 0)}%"),
            ]
            
            for i, (metric, value) in enumerate(metrics_data, start=metrics_row + 1):
                ws[f"A{i}"] = metric
                ws[f"B{i}"] = value
                
            # Add a simple bar chart
            if len(metrics_data) > 0:
                chart = BarChart()
                chart.title = "Key Metrics"
                chart.style = 10
                chart.y_axis.title = "Count"
                chart.x_axis.title = "Metric"
                
                # Data for chart (excluding percentage values)
                numeric_metrics = [(m, v) for m, v in metrics_data if isinstance(v, (int, float))]
                if numeric_metrics:
                    data = Reference(ws, min_col=2, min_row=metrics_row+1, max_row=metrics_row+len(numeric_metrics))
                    categories = Reference(ws, min_col=1, min_row=metrics_row+1, max_row=metrics_row+len(numeric_metrics))
                    chart.add_data(data, titles_from_data=False)
                    chart.set_categories(categories)
                    ws.add_chart(chart, "D3")
                    
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
            
    async def _create_phishing_excel(self, wb: Workbook, phishing_data: Dict[str, Any]):
        """Create phishing statistics sheet in Excel."""
        ws = wb.create_sheet("Phishing Statistics")
        
        # Summary section
        if "summary" in phishing_data:
            summary = phishing_data["summary"]
            
            ws["A1"] = "Phishing Campaign Summary"
            ws["A1"].font = Font(bold=True, size=16)
            
            row = 3
            ws[f"A{row}"] = "Total Sent"
            ws[f"B{row}"] = summary.get("total_sent", 0)
            
            row += 1
            ws[f"A{row}"] = "Opened"
            ws[f"B{row}"] = summary.get("opened", 0)
            
            row += 1
            ws[f"A{row}"] = "Clicked"
            ws[f"B{row}"] = summary.get("clicked", 0)
            
            row += 1
            ws[f"A{row}"] = "Reported"
            ws[f"B{row}"] = summary.get("reported", 0)
            
            # Add pie chart for phishing results
            pie = PieChart()
            labels = Reference(ws, min_col=1, min_row=4, max_row=6)
            data = Reference(ws, min_col=2, min_row=4, max_row=6)
            pie.add_data(data)
            pie.set_categories(labels)
            pie.title = "Phishing Results Distribution"
            ws.add_chart(pie, "D3")
            
    # PDF Helper Methods
    
    async def _create_executive_summary(self, overview: Dict[str, Any]) -> List:
        """Create executive summary section for PDF."""
        elements = []
        
        elements.append(Paragraph("Executive Summary", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Key metrics table
        data = [
            ["Metric", "Value", "Status"],
            ["Total Users", str(overview.get("total_users", 0)), "Active"],
            ["Engagement Rate", f"{overview.get('engagement_rate', 0)}%", self._get_status_indicator(overview.get('engagement_rate', 0), 70, 50)],
            ["Training Completion", f"{overview.get('courses_completed', 0)}", "Ongoing"],
            ["Average Score", f"{overview.get('avg_quiz_score', 0)}%", self._get_status_indicator(overview.get('avg_quiz_score', 0), 80, 60)],
        ]
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        return elements
        
    async def _create_security_score_section(self, security_data: Dict[str, Any]) -> List:
        """Create security score section for PDF."""
        elements = []
        
        elements.append(Paragraph("Security Score Analysis", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Overall score
        score = security_data.get("score", 0)
        score_color = self._get_score_color(score)
        
        elements.append(Paragraph(
            f"Overall Security Score: <font color='{score_color}'>{score}/100</font>",
            self.styles['Heading3']
        ))
        elements.append(Spacer(1, 0.1 * inch))
        
        # Breakdown table
        if "breakdown" in security_data:
            breakdown = security_data["breakdown"]
            data = [["Component", "Score", "Weight"]]
            
            for component, value in breakdown.items():
                data.append([
                    component.replace("_", " ").title(),
                    f"{value}",
                    "25%"  # Example weight
                ])
                
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            
        # Recommendations
        if "recommendations" in security_data:
            elements.append(Spacer(1, 0.2 * inch))
            elements.append(Paragraph("Recommendations:", self.styles['Heading4']))
            
            for rec in security_data["recommendations"]:
                elements.append(Paragraph(f"• {rec}", self.styles['Normal']))
                
        return elements
        
    def _get_status_indicator(self, value: float, good_threshold: float, warning_threshold: float) -> str:
        """Get status indicator based on thresholds."""
        if value >= good_threshold:
            return "Good"
        elif value >= warning_threshold:
            return "Warning"
        else:
            return "Critical"
            
    def _get_score_color(self, score: float) -> str:
        """Get color based on score."""
        if score >= 80:
            return "green"
        elif score >= 60:
            return "orange"
        else:
            return "red"